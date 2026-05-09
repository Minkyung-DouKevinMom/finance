"""
우리집 금융 관리 앱 (Streamlit)
실행: streamlit run app.py
"""

from datetime import date, datetime
from dateutil.relativedelta import relativedelta

import pandas as pd
import streamlit as st

import repository as repo
from db import init_db

# ─────────────────────── 초기화 ────────────────────────
init_db()

st.set_page_config(page_title="우리집 금융 관리", page_icon="💰", layout="wide")

# ─────────────────────── 사이드바 ──────────────────────
PAGES = {
    "📊 대시보드": "dashboard",
    "📅 월별 스냅샷": "snapshot",
    "💼 계좌(상품) 관리": "accounts",
    "👴 연금 시뮬레이션": "pension",
    "📈 자산 추이": "trend",
    "🎯 미래 계획": "plan",
    "📥 엑셀 Import": "import_excel",
    "⚙️ 기준정보 관리": "master",
}
choice = st.sidebar.radio("메뉴", list(PAGES.keys()))
page = PAGES[choice]

st.sidebar.markdown("---")
st.sidebar.caption("금액 단위: **만원**")


# ─────────────────────── 공통 유틸 ─────────────────────
def fmt_won(v):
    """만원 단위 숫자를 천 단위 콤마로 표기."""
    if v is None or pd.isna(v):
        return "-"
    return f"{int(round(v)):,}"


def parse_date_safely(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


# =========================================================
# 1) 대시보드
# =========================================================
def page_dashboard():
    st.title("📊 대시보드")
    snapshots = repo.list_snapshots()
    if not snapshots:
        st.info("아직 스냅샷이 없습니다. 좌측 메뉴 **월별 스냅샷**에서 첫 데이터를 입력해 주세요.")
        return

    sel = st.selectbox(
        "기준 시점",
        options=snapshots,
        format_func=lambda s: s["snapshot_date"][:7],
    )
    summary = repo.snapshot_summary(sel["id"])
    if summary.empty:
        st.warning("선택한 시점에 입력된 잔액이 없습니다.")
        return

    total = summary["total"].sum()

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("총자산 (만원)", fmt_won(total))
    for col, cat_name in zip(
        [col2, col3, col4], ["금융자산", "부동산", "실물자산"]
    ):
        v = summary.loc[summary["category"] == cat_name, "total"].sum()
        ratio = v / total * 100 if total else 0
        col.metric(f"{cat_name}", fmt_won(v), f"{ratio:.1f}%")

    st.markdown("### 분류별 비중")
    cat_sum = summary.groupby("category", as_index=False)["total"].sum()
    st.bar_chart(cat_sum.set_index("category"))

    st.markdown("### 과목별 상세")
    detail = summary.copy()
    detail["비율(%)"] = (detail["total"] / total * 100).round(1)
    detail = detail.rename(columns={
        "category": "자산구분", "subcategory": "과목", "total": "금액(만원)"
    })
    detail["금액(만원)"] = detail["금액(만원)"].apply(fmt_won)
    st.dataframe(detail, use_container_width=True, hide_index=True)


# =========================================================
# 2) 월별 스냅샷 (잔액 입력/수정)
# =========================================================
def page_snapshot():
    st.title("📅 월별 스냅샷")

    snapshots = repo.list_snapshots()
    options = ["+ 새 시점 만들기"] + [s["snapshot_date"][:7] for s in snapshots]
    pick = st.selectbox("시점 선택", options)

    if pick == "+ 새 시점 만들기":
        with st.form("new_snap"):
            d = st.date_input("스냅샷 날짜", value=date.today().replace(day=1))
            memo = st.text_input("메모 (선택)")
            submitted = st.form_submit_button("생성")
            if submitted:
                d = d.replace(day=1)
                repo.get_or_create_snapshot(d.isoformat(), memo)
                st.success(f"{d.strftime('%Y-%m')} 스냅샷 준비 완료. 다시 선택해 주세요.")
                st.rerun()
        return

    snap = next(s for s in snapshots if s["snapshot_date"][:7] == pick)

    df = repo.get_balances_for_snapshot(snap["id"])
    if df.empty:
        st.warning("등록된 활성 계좌가 없습니다. 먼저 **계좌 관리**에서 상품을 등록해 주세요.")
        return

    tab_edit, tab_delete = st.tabs(["✏️ 금액 수정", "🗑️ 항목 삭제"])

    # ── 탭 1: 금액 수정 ──────────────────────────────────
    with tab_edit:
        st.caption("표 안에서 **금액(만원)** 셀을 직접 클릭해 수정한 뒤 **저장** 버튼을 눌러 주세요.")

        edit_df = df[[
            "category_name", "subcategory_name", "owner_name",
            "account_name", "product_name", "status", "amount", "account_id"
        ]].rename(columns={
            "category_name": "자산구분",
            "subcategory_name": "과목",
            "owner_name": "소유자",
            "account_name": "항목",
            "product_name": "상품명",
            "status": "납입상태",
            "amount": "금액(만원)",
        })

        edited = st.data_editor(
            edit_df,
            hide_index=True,
            use_container_width=True,
            disabled=["자산구분", "과목", "소유자", "항목", "상품명", "납입상태", "account_id"],
            column_config={
                "account_id": None,
                "금액(만원)": st.column_config.NumberColumn(format="%d", min_value=-999999),
            },
            key=f"editor_{snap['id']}",
        )

        # 저장 / 스냅샷 삭제 버튼
        c1, c2, c3 = st.columns([1, 1, 4])
        if c1.button("💾 저장", type="primary"):
            repo.save_balances(
                snap["id"],
                [{"account_id": int(r["account_id"]), "amount": float(r["금액(만원)"] or 0)}
                 for _, r in edited.iterrows()],
            )
            st.success("저장되었습니다.")
            st.rerun()

        if c2.button("🗑️ 이 스냅샷 전체 삭제"):
            repo.delete_snapshot(snap["id"])
            st.success("스냅샷이 삭제되었습니다.")
            st.rerun()

        # 합계 요약
        st.markdown("---")
        st.markdown("#### 합계")
        by_cat = edited.groupby("자산구분", as_index=False)["금액(만원)"].sum()
        by_cat["금액(만원)"] = by_cat["금액(만원)"].apply(fmt_won)
        by_sub = edited.groupby(["자산구분", "과목"], as_index=False)["금액(만원)"].sum()
        by_sub["금액(만원)"] = by_sub["금액(만원)"].apply(fmt_won)

        cc1, cc2 = st.columns(2)
        with cc1:
            st.write("**자산구분 합계**")
            st.dataframe(by_cat, hide_index=True, use_container_width=True)
        with cc2:
            st.write("**과목 합계**")
            st.dataframe(by_sub, hide_index=True, use_container_width=True)

        st.metric("총자산 (만원)", fmt_won(edited["금액(만원)"].sum()))

    # ── 탭 2: 개별 항목 삭제 ────────────────────────────
    with tab_delete:
        st.caption("이 시점에서 **특정 항목의 잔액 기록만** 삭제합니다. 계좌 마스터는 유지됩니다.")

        # 잔액이 실제로 기록된 항목만 표시 (amount > 0 또는 balance 레코드 존재)
        has_balance_df = df[df["amount"] != 0].copy()
        if has_balance_df.empty:
            st.info("삭제할 잔액 기록이 없습니다.")
        else:
            # 삭제 대상 선택 (멀티셀렉트)
            options_map = {
                f"{r['owner_name']} | {r['category_name']} > {r['subcategory_name']} | {r['account_name']} ({int(r['amount']):,}만원)": int(r["account_id"])
                for _, r in has_balance_df.iterrows()
            }
            selected_labels = st.multiselect(
                "삭제할 항목 선택 (복수 선택 가능)",
                options=list(options_map.keys()),
            )

            if selected_labels:
                st.warning(f"선택된 {len(selected_labels)}개 항목의 **이 시점 잔액 기록**이 삭제됩니다.")
                if st.button("🗑️ 선택 항목 삭제", type="primary"):
                    selected_ids = [options_map[l] for l in selected_labels]
                    repo.delete_balances(snap["id"], selected_ids)
                    st.success(f"{len(selected_ids)}개 항목 삭제 완료.")
                    st.rerun()

        st.markdown("---")
        st.markdown("#### 전체 항목 목록 (현재 시점)")
        view_df = df[[
            "owner_name", "category_name", "subcategory_name", "account_name", "amount"
        ]].rename(columns={
            "owner_name": "소유자", "category_name": "자산구분",
            "subcategory_name": "과목", "account_name": "항목", "amount": "금액(만원)"
        }).copy()
        view_df["금액(만원)"] = view_df["금액(만원)"].apply(fmt_won)
        st.dataframe(view_df, hide_index=True, use_container_width=True)


# =========================================================
# 3) 계좌(상품) 관리
# =========================================================
def page_accounts():
    st.title("💼 계좌(상품) 관리")
    st.caption("개시일·상품명 등 한 번 등록한 정보는 이후 자동으로 사용됩니다.")

    tab_list, tab_add = st.tabs(["📋 목록", "➕ 추가"])

    # ----- 목록 / 수정 -----
    with tab_list:
        show_inactive = st.checkbox("비활성 항목 포함", value=False)
        accounts = repo.list_accounts(active_only=not show_inactive)
        if not accounts:
            st.info("등록된 계좌가 없습니다.")
        else:
            df = pd.DataFrame(accounts)[[
                "id", "category_name", "subcategory_name", "owner_name",
                "name", "product_name", "start_date", "payout_start_date",
                "monthly_premium", "expected_payout", "status", "is_active",
            ]].rename(columns={
                "id": "ID", "category_name": "자산구분", "subcategory_name": "과목",
                "owner_name": "소유자", "name": "항목", "product_name": "상품명",
                "start_date": "시작일", "payout_start_date": "개시일",
                "monthly_premium": "납입(만원)", "expected_payout": "예상수령(만원)",
                "status": "상태", "is_active": "활성",
            })
            st.dataframe(df, hide_index=True, use_container_width=True)

        st.markdown("---")
        st.markdown("#### 계좌 수정 / 삭제")
        if accounts:
            opts = {f"#{a['id']} {a['name']} ({a['owner_name']})": a["id"]
                    for a in accounts}
            label = st.selectbox("수정할 계좌", list(opts.keys()))
            account_id = opts[label]
            _account_form(account_id=account_id)

    # ----- 추가 -----
    with tab_add:
        _account_form(account_id=None)


def _account_form(account_id=None):
    """계좌 등록/수정 공용 폼."""
    owners = repo.list_owners()
    subs = repo.list_subcategories()

    existing = repo.get_account(account_id) if account_id else None

    with st.form(f"acc_form_{account_id or 'new'}", clear_on_submit=(account_id is None)):
        c1, c2, c3 = st.columns(3)
        sub_id = c1.selectbox(
            "과목",
            options=[s["id"] for s in subs],
            format_func=lambda x: next(
                f"{s['category_name']} > {s['name']}" for s in subs if s["id"] == x),
            index=([s["id"] for s in subs].index(existing["subcategory_id"])
                   if existing else 0),
        )
        owner_id = c2.selectbox(
            "소유자",
            options=[o["id"] for o in owners],
            format_func=lambda x: next(o["name"] for o in owners if o["id"] == x),
            index=([o["id"] for o in owners].index(existing["owner_id"])
                   if existing else 0),
        )
        is_active = c3.checkbox("활성", value=bool(existing["is_active"]) if existing else True)

        name = st.text_input("항목명*", value=existing["name"] if existing else "")
        product_name = st.text_input("상품명",
                                     value=existing["product_name"] if existing else "")

        d1, d2, d3 = st.columns(3)
        DATE_MIN = date(1980, 1, 1)
        DATE_MAX = date(date.today().year + 50, 12, 31)
        start_date = d1.date_input(
            "시작일",
            value=parse_date_safely(existing["start_date"] if existing else None),
            min_value=DATE_MIN, max_value=DATE_MAX)
        maturity_date = d2.date_input(
            "만기일",
            value=parse_date_safely(existing["maturity_date"] if existing else None),
            min_value=DATE_MIN, max_value=DATE_MAX)
        payout_start_date = d3.date_input(
            "연금 개시일",
            value=parse_date_safely(existing["payout_start_date"] if existing else None),
            min_value=DATE_MIN, max_value=DATE_MAX,
            help="연금 수령이 시작되는 날짜. 연금 시뮬레이션에 사용됩니다.",
        )

        m1, m2, m3 = st.columns(3)
        monthly_premium = m1.number_input(
            "누적 납입(만원)",
            value=float(existing["monthly_premium"]) if existing and existing["monthly_premium"] else 0.0,
            step=100.0,
        )
        expected_payout = m2.number_input(
            "예상 수령액(만원)",
            value=float(existing["expected_payout"]) if existing and existing["expected_payout"] else 0.0,
            step=100.0,
            help="연금: 월/연 수령액 / 일시불: 총액",
        )
        PAYOUT_OPTIONS = ["monthly", "yearly", "lumpsum"]
        PAYOUT_LABELS  = {"monthly": "월 수령", "yearly": "연 수령", "lumpsum": "일시불"}
        cur_payout = existing["payout_type"] if existing and existing["payout_type"] in PAYOUT_OPTIONS else "monthly"
        payout_type = m3.selectbox(
            "수령 형태",
            options=PAYOUT_OPTIONS,
            format_func=lambda x: PAYOUT_LABELS.get(x, x),
            index=PAYOUT_OPTIONS.index(cur_payout),
        )

        status = st.selectbox(
            "납입 상태",
            ["", "납입중", "납입완료", "해지"],
            index=(["", "납입중", "납입완료", "해지"].index(existing["status"])
                   if existing and existing["status"] in ["", "납입중", "납입완료", "해지"]
                   else 0),
        )
        memo = st.text_area("메모", value=existing["memo"] if existing and existing["memo"] else "")

        col_save, col_del = st.columns([1, 1])
        save = col_save.form_submit_button("💾 저장", type="primary")
        do_delete = col_del.form_submit_button("🗑️ 삭제") if account_id else False

        if save:
            if not name:
                st.error("항목명은 필수입니다.")
            else:
                repo.upsert_account(
                    account_id=account_id,
                    subcategory_id=sub_id,
                    owner_id=owner_id,
                    name=name,
                    product_name=product_name or None,
                    start_date=start_date.isoformat() if start_date else None,
                    maturity_date=maturity_date.isoformat() if maturity_date else None,
                    payout_start_date=payout_start_date.isoformat() if payout_start_date else None,
                    monthly_premium=monthly_premium or None,
                    expected_payout=expected_payout or None,
                    payout_type=payout_type,
                    status=status or None,
                    memo=memo or None,
                    is_active=1 if is_active else 0,
                )
                st.success("저장되었습니다.")
                st.rerun()

        if do_delete:
            result = repo.delete_account(account_id)
            if result == "deactivated":
                st.warning("잔액 기록이 있어 비활성화 처리했습니다.")
            else:
                st.success("삭제되었습니다.")
            st.rerun()


# =========================================================
# 4) 연금 시뮬레이션
# =========================================================
def page_pension():
    st.title("👴 연금 시뮬레이션")
    st.caption("개시일을 기준으로 향후 월별 수령 합계를 계산합니다.")

    accounts = repo.list_accounts(active_only=True)
    pension_accounts = [
        a for a in accounts
        if a["expected_payout"] and a["expected_payout"] > 0
    ]
    if not pension_accounts:
        st.info("예상 수령액이 입력된 계좌가 없습니다. **계좌 관리**에서 입력해 주세요.")
        return

    today = date.today()
    target = st.slider(
        "조회 시점 (년)", min_value=today.year, max_value=today.year + 40,
        value=today.year + 10, step=1,
    )
    target_date = date(target, today.month, 1)

    rows = []
    monthly_total = 0.0
    yearly_total  = 0.0
    lumpsum_total = 0.0
    PAYOUT_LABELS = {"monthly": "월 수령", "yearly": "연 수령", "lumpsum": "일시불"}
    for a in pension_accounts:
        payout_d = parse_date_safely(a["payout_start_date"])
        started = payout_d is not None and payout_d <= target_date
        rows.append({
            "소유자": a["owner_name"],
            "항목": a["name"],
            "상품명": a["product_name"] or "",
            "개시일": a["payout_start_date"] or "-",
            "수령형태": PAYOUT_LABELS.get(a["payout_type"], a["payout_type"]),
            "예상수령액(만원)": a["expected_payout"],
            "수령시작?": "✅" if started else "—",
        })
        if started:
            if a["payout_type"] == "monthly":
                monthly_total += a["expected_payout"]
            elif a["payout_type"] == "yearly":
                yearly_total += a["expected_payout"]
            else:
                lumpsum_total += a["expected_payout"]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("조회 시점", target_date.strftime("%Y-%m"))
    c2.metric("월 수령 합계 (만원)", fmt_won(monthly_total))
    c3.metric("연 수령 합계 (만원)", fmt_won(yearly_total),
              help="연 수령액 ÷ 12 = 월 환산 " + fmt_won(yearly_total / 12 if yearly_total else 0) + "만원")
    c4.metric("일시불 누적 (만원)", fmt_won(lumpsum_total))

    df = pd.DataFrame(rows)
    df["예상수령액(만원)"] = df["예상수령액(만원)"].apply(fmt_won)
    st.dataframe(df, hide_index=True, use_container_width=True)

    # 연도별 월 환산 수령 추이 (연 수령은 ÷12 해서 합산)
    st.markdown("---")
    st.markdown("### 연도별 월 환산 수령 추이")
    st.caption("연 수령 상품은 ÷ 12 하여 월 환산 합계에 포함합니다.")
    timeline = []
    for y in range(today.year, today.year + 41):
        d = date(y, 1, 1)
        m_sum = 0.0
        for a in pension_accounts:
            payout_d = parse_date_safely(a["payout_start_date"])
            if not payout_d or payout_d > d:
                continue
            if a["payout_type"] == "monthly":
                m_sum += a["expected_payout"]
            elif a["payout_type"] == "yearly":
                m_sum += a["expected_payout"] / 12
        timeline.append({"연도": y, "월 환산 수령(만원)": round(m_sum, 1)})
    tdf = pd.DataFrame(timeline).set_index("연도")
    st.line_chart(tdf)


# =========================================================
# 5) 자산 추이
# =========================================================
def page_trend():
    st.title("📈 자산 추이")
    df = repo.time_series_total()
    if df.empty:
        st.info("표시할 데이터가 없습니다.")
        return

    df["date"] = pd.to_datetime(df["date"]).dt.strftime("%Y-%m")
    pivot = df.pivot(index="date", columns="category", values="total").fillna(0)
    pivot["전체"] = pivot.sum(axis=1)

    st.markdown("### 카테고리별 추이")
    st.line_chart(pivot)

    st.markdown("### 시점별 합계 표")
    show = pivot.copy().applymap(fmt_won)
    st.dataframe(show, use_container_width=True)


# =========================================================
# 6) 미래 계획
# =========================================================
def page_plan():
    st.title("🎯 미래 계획")
    plans = repo.list_plans()

    if not plans.empty:
        view = plans.rename(columns={
            "id": "ID", "plan_year": "연도", "category": "구분",
            "item": "항목", "target_amount": "금액(만원)", "age": "나이", "memo": "비고",
        })
        st.dataframe(view, hide_index=True, use_container_width=True)

    st.markdown("---")
    st.markdown("#### 추가 / 수정")
    plan_id = None
    if not plans.empty:
        opts = {f"#{r['id']} {r['item']} ({r['plan_year']})": r["id"]
                for _, r in plans.iterrows()}
        opts["+ 새 항목 추가"] = None
        label = st.selectbox("선택", list(opts.keys()))
        plan_id = opts[label]

    existing = None
    if plan_id is not None:
        existing = plans[plans["id"] == plan_id].iloc[0].to_dict()

    with st.form(f"plan_form_{plan_id or 'new'}"):
        c1, c2, c3 = st.columns(3)
        plan_year = c1.number_input(
            "연도", min_value=2000, max_value=2100,
            value=int(existing["plan_year"]) if existing and existing["plan_year"] else date.today().year,
        )
        category = c2.text_input("구분", value=existing["category"] if existing else "")
        age = c3.number_input(
            "나이", min_value=0, max_value=120,
            value=int(existing["age"]) if existing and existing["age"] else 0,
        )
        item = st.text_input("항목*", value=existing["item"] if existing else "")
        amount = st.number_input(
            "금액(만원)", step=100.0,
            value=float(existing["target_amount"]) if existing and existing["target_amount"] else 0.0,
        )
        memo = st.text_area("비고", value=existing["memo"] if existing and existing["memo"] else "")

        col_save, col_del = st.columns([1, 1])
        save = col_save.form_submit_button("💾 저장", type="primary")
        do_delete = col_del.form_submit_button("🗑️ 삭제") if plan_id else False

        if save:
            if not item:
                st.error("항목은 필수입니다.")
            else:
                repo.upsert_plan(
                    plan_id=plan_id, plan_year=int(plan_year),
                    category=category or None, item=item,
                    target_amount=amount or None,
                    age=int(age) if age else None, memo=memo or None,
                )
                st.success("저장되었습니다.")
                st.rerun()

        if do_delete:
            repo.delete_plan(plan_id)
            st.success("삭제되었습니다.")
            st.rerun()


# =========================================================
# 7) 기준정보 관리
# =========================================================
def page_master():
    st.title("⚙️ 기준정보 관리")

    tab_owner, tab_cat, tab_sub = st.tabs(["소유자", "자산구분", "과목"])

    with tab_owner:
        st.markdown("#### 소유자 목록")
        st.dataframe(pd.DataFrame(repo.list_owners()),
                     hide_index=True, use_container_width=True)
        with st.form("add_owner"):
            n = st.text_input("새 소유자")
            if st.form_submit_button("추가") and n:
                repo.add_owner(n)
                st.rerun()

    with tab_cat:
        st.markdown("#### 자산구분 (대분류)")
        st.dataframe(pd.DataFrame(repo.list_categories()),
                     hide_index=True, use_container_width=True)
        with st.form("add_cat"):
            n = st.text_input("새 자산구분")
            o = st.number_input("정렬", min_value=0, max_value=999, value=99)
            if st.form_submit_button("추가") and n:
                repo.add_category(n, int(o))
                st.rerun()

    with tab_sub:
        st.markdown("#### 과목 (중분류)")
        subs = repo.list_subcategories()
        st.dataframe(pd.DataFrame(subs), hide_index=True, use_container_width=True)
        cats = repo.list_categories()
        with st.form("add_sub"):
            cid = st.selectbox(
                "자산구분",
                options=[c["id"] for c in cats],
                format_func=lambda x: next(c["name"] for c in cats if c["id"] == x),
            )
            n = st.text_input("새 과목")
            o = st.number_input("정렬", min_value=0, max_value=999, value=99, key="sub_order")
            if st.form_submit_button("추가") and n:
                repo.add_subcategory(cid, n, int(o))
                st.rerun()


# =========================================================
# 8) 엑셀 Import
# =========================================================
def page_import_excel():
    from importer import import_workbook
    import io

    st.title("📥 엑셀 일괄 Import")
    st.caption(
        "기존 `금융현황_YYYYMM` 형식의 시트가 담긴 엑셀 파일을 업로드하면 "
        "모든 시점을 한 번에 DB에 가져옵니다."
    )

    st.info(
        "**처리 방식**\n"
        "- 계좌(상품) 마스터가 없으면 자동 신규 등록, 있으면 메타정보만 보완합니다.\n"
        "- 잔액은 **덮어쓰기(upsert)** 방식으로 저장됩니다 (이미 입력된 시점도 갱신).\n"
        "- `금융현황_YYYYMM` 패턴과 일치하지 않는 시트는 무시합니다.",
        icon="ℹ️",
    )

    uploaded = st.file_uploader("엑셀 파일 선택 (.xlsx)", type=["xlsx"])

    if not uploaded:
        return

    # 파일 미리보기 (시트 목록)
    from openpyxl import load_workbook
    import re
    buf = uploaded.read()
    wb_preview = load_workbook(io.BytesIO(buf), data_only=True, read_only=True)
    target_sheets = sorted(
        [n for n in wb_preview.sheetnames if re.match(r"금융현황_\d{6}$", n)],
        key=lambda n: n.split("_")[1],
    )
    wb_preview.close()

    if not target_sheets:
        st.error("파일 내에 `금융현황_YYYYMM` 형식의 시트가 없습니다.")
        return

    st.markdown(f"**감지된 시트 {len(target_sheets)}개:** " + ", ".join(target_sheets))

    if not st.button("🚀 Import 시작", type="primary"):
        return

    progress = st.progress(0, text="Import 준비 중...")
    result = import_workbook(io.BytesIO(buf))
    sheets = result["sheets"]

    # 결과 집계
    total_imported = sum(s["imported"] for s in sheets)
    total_errors   = sum(len(s["errors"]) for s in sheets)
    progress.progress(1.0, text="완료!")

    # 요약 메트릭
    c1, c2, c3 = st.columns(3)
    c1.metric("처리한 시트", len(sheets))
    c2.metric("저장된 잔액 건수", total_imported)
    c3.metric("오류", total_errors, delta_color="inverse")

    # 시트별 상세 결과
    st.markdown("---")
    st.markdown("#### 시트별 처리 결과")

    for s in sheets:
        has_error = len(s["errors"]) > 0
        icon = "✅" if not has_error else "⚠️"
        label = (
            f"{icon} **{s['name']}**  →  {s['snapshot_date']}  |  "
            f"파싱 {len(s['records'])}건  /  저장 {s['imported']}건"
            + (f"  /  오류 {len(s['errors'])}건" if has_error else "")
        )
        with st.expander(label, expanded=has_error):
            if s["errors"]:
                st.error("\n".join(s["errors"]))
            if s["records"]:
                df = (
                    pd.DataFrame(s["records"])[
                        ["category", "subcategory", "owner", "account_name", "amount"]
                    ]
                    .rename(columns={
                        "category": "자산구분", "subcategory": "과목",
                        "owner": "소유자", "account_name": "항목", "amount": "금액(만원)",
                    })
                )
                df["금액(만원)"] = df["금액(만원)"].apply(fmt_won)
                st.dataframe(df, hide_index=True, use_container_width=True)

    if total_errors == 0:
        st.success("모든 시트를 성공적으로 Import했습니다. 대시보드에서 확인해 보세요!")
    else:
        st.warning("일부 항목에 오류가 발생했습니다. 위 내용을 확인 후 수동으로 보완해 주세요.")


# ─────────────────────── 라우팅 ────────────────────────
PAGE_FUNCS = {
    "dashboard":     page_dashboard,
    "snapshot":      page_snapshot,
    "accounts":      page_accounts,
    "pension":       page_pension,
    "trend":         page_trend,
    "plan":          page_plan,
    "import_excel":  page_import_excel,
    "master":        page_master,
}
PAGE_FUNCS[page]()
