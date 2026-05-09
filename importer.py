"""
엑셀 금융현황_YYYYMM 시트 일괄 Import.

지원 레이아웃:
  - Layout A: 202108~현재  헤더 row0 = ('자산', None, '과목', '시작일', '만기일', '개시일', '비율', '금액', ...)
  - Layout B: 201910       헤더 row0 = (None, '과목', '시작일', '만기일', '개시일', '월납부금액', '금액', '금액(-)', '합계', ...)
"""

import re
from datetime import date
from openpyxl import load_workbook

import repository as repo
from db import get_conn

# ─────────────────────── 정규화 테이블 ───────────────────────

# 엑셀 과목(소분류) → DB 서브카테고리명으로 정규화
SUBCATEGORY_NORMALIZE = {
    # 보험·연금 계열
    "보험,연금": "보험·연금",
    "보험,연금 (사용불가/노후자금)": "보험·연금",
    "보험,연금\n(사용불가/노후자금)": "보험·연금",
    "적금": "보험·연금",
    # 예금·대출 계열
    "예금": "예금·대출",
    "예금,대출": "예금·대출",
    "대출": "예금·대출",
    "저축": "예금·대출",
    # 주식·펀드 계열
    "주식,펀드": "주식·펀드",
    "주식": "주식·펀드",
    "펀드": "주식·펀드",
}

# 대분류 → DB 카테고리명
CATEGORY_NORMALIZE = {
    "금융자산": "금융자산",
    "적금": "금융자산",
    "대출": "금융자산",
    "부동산": "부동산",
    "실물자산": "실물자산",
}

# 계정명에 "(오빠)" 포함 → 배우자
def infer_owner(account_name: str) -> str:
    if "(오빠)" in account_name or "(남편)" in account_name:
        return "배우자"
    return "본인"

# 금액 없는 행이나 skip 대상 여부
SKIP_NAMES = {"합계", "전체", "과목", "항목", "구분", "전체항목"}

def is_skip_row(name):
    if not name:
        return True
    name_s = str(name).strip()
    if not name_s:
        return True
    if name_s in SKIP_NAMES:
        return True
    # 연도 계획 섹션 헤더
    if re.match(r"20\d\d년", name_s):
        return True
    # 날짜처럼 생긴 이름 (계획 섹션 시기 컬럼이 항목명 컬럼에 들어온 경우)
    if re.match(r"20\d\d[.\-/]\d{1,2}", name_s):
        return True
    # "급여개시일" 같은 헤더
    if name_s in ("급여개시일", "납입보험료", "예상금액", "비고", "시기", "나이"):
        return True
    return False

def parse_date_str(v) -> str | None:
    """엑셀에서 온 날짜 값을 YYYY-MM-DD 문자열로 변환. 실패 시 None."""
    if v is None:
        return None
    s = str(v).strip()
    # 2007.12.21 형식
    m = re.match(r"(\d{4})\.(\d{1,2})\.(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    # 2017.06 형식 (float → str 됨)
    m = re.match(r"(\d{4})\.(\d{1,2})$", s)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-01"
    # 2017 (년도만, int)
    m = re.match(r"^(20\d{2})$", s)
    if m:
        return f"{m.group(1)}-01-01"
    return None

def sheet_name_to_date(sheet_name: str) -> str | None:
    """금융현황_202601 → 2026-01-01"""
    m = re.search(r"(\d{4})(\d{2})$", sheet_name)
    if m:
        return f"{m.group(1)}-{m.group(2)}-01"
    return None


# ─────────────────────── 레이아웃 감지 ───────────────────────

def detect_layout(header_row) -> str:
    """
    'A': 자산 컬럼 위치 0, 과목 위치 2, 금액 위치 7
    'B': 과목 위치 1, 금액 위치 6  (201910 형태)
    """
    if header_row[0] == "자산":
        return "A"
    if header_row[1] == "과목":
        return "B"
    return "A"  # 기본값


# ─────────────────────── 시트 파싱 ───────────────────────────

def parse_sheet(ws) -> list[dict]:
    """
    한 시트에서 개별 금융 항목 리스트를 추출.
    반환: [{category, subcategory, account_name, amount, start_date, maturity_date,
            payout_start_date, status, product_name, memo, owner}, ...]
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    layout = detect_layout(rows[0])

    if layout == "A":
        IDX_ASSET = 0   # 자산 대분류
        IDX_SUB   = 1   # 과목(소분류)
        IDX_NAME  = 2   # 항목명
        IDX_START = 3
        IDX_MATUR = 4
        IDX_PAYOUT= 5
        IDX_AMT   = 7
        IDX_STATUS= 8
        IDX_PROD  = 9   # 상품명 (없는 버전도 있음)
        IDX_MEMO  = 10
    else:  # layout B (201910)
        # 컬럼: [0]=대분류, [1]=과목/항목, [2]=시작일, [3]=만기일, [4]=개시일,
        #        [5]=월납부금액, [6]=금액(자산), [7]=금액(-)(부채), [8]=합계, [9]=기타(국민연금금액/상태)
        IDX_ASSET = 0
        IDX_SUB   = 0   # 대분류와 같은 컬럼이 과목 역할도 함
        IDX_NAME  = 1
        IDX_START = 2
        IDX_MATUR = 3
        IDX_PAYOUT= 4
        IDX_AMT   = 6   # 기본 금액 컬럼
        IDX_STATUS= 9
        IDX_PROD  = None
        IDX_MEMO  = None

    records = []
    cur_category = None      # 대분류 (금융자산/부동산/실물자산)
    cur_subcategory = None   # 과목 (보험·연금/예금·대출/주식·펀드)
    in_asset_section = False # 자산 섹션 내부인지

    def safe(row, idx):
        try:
            return row[idx]
        except (IndexError, TypeError):
            return None

    for row in rows[1:]:
        row = list(row)

        # ── 대분류 감지 ──
        v_asset = safe(row, IDX_ASSET) if IDX_ASSET is not None else None
        if v_asset in CATEGORY_NORMALIZE:
            cur_category = CATEGORY_NORMALIZE[v_asset]
            in_asset_section = True
            # Layout B: 대분류가 과목을 겸함 (적금→보험·연금, 부동산→부동산, 대출→예금·대출)
            if layout == "B":
                sub_from_cat = {
                    "적금": "보험·연금", "부동산": "부동산",
                    "대출": "예금·대출", "금융자산": "보험·연금"
                }
                if str(v_asset) in sub_from_cat:
                    cur_subcategory = sub_from_cat[str(v_asset)]

        # 자산 섹션 종료 신호 (전체합계 행, 연도계획 헤더)
        v0 = safe(row, 0)
        if v0 == "전체":
            in_asset_section = False
        if isinstance(v0, str) and re.match(r"20\d\d년", v0.strip()):
            in_asset_section = False
        if not in_asset_section and layout == "A":
            continue

        # ── 과목(소분류) 감지 ──
        v_sub = safe(row, IDX_SUB)
        if v_sub and str(v_sub).strip() not in SKIP_NAMES:
            normalized = SUBCATEGORY_NORMALIZE.get(
                str(v_sub).strip().replace("\n", "\n"),  # 원본 그대로
                SUBCATEGORY_NORMALIZE.get(str(v_sub).strip()))
            if normalized:
                cur_subcategory = normalized
            elif layout == "A" and str(v_sub).strip() not in ("합계",):
                # 과목이 정규화 안 되면 그냥 원본으로 (임포트 후 수동 매핑 가능)
                pass

        # ── 항목명 ──
        v_name = safe(row, IDX_NAME)
        if is_skip_row(v_name):
            continue

        # ── 금액 ──
        v_amt = safe(row, IDX_AMT)
        # Layout B: 금액(idx=6) 없으면 기타(idx=9, 국민연금 케이스), 부채는 금액(-)(idx=7)을 음수로
        if layout == "B":
            v_amt_asset  = safe(row, 6)   # 금액(자산)
            v_amt_debt   = safe(row, 7)   # 금액(-)(부채)
            v_amt_extra  = safe(row, 9)   # 기타(국민연금 금액 등)
            if v_amt_asset is not None:
                v_amt = float(v_amt_asset)
            elif v_amt_debt is not None:
                try:
                    v_amt = -abs(float(v_amt_debt))
                except (TypeError, ValueError):
                    v_amt = 0.0
            elif isinstance(v_amt_extra, (int, float)):
                # 기타 컬럼에 숫자 → 국민연금처럼 금액이 여기 들어오는 케이스
                v_amt = float(v_amt_extra)
            else:
                v_amt = 0.0
            # 대출 대분류 행: 금액(-) 컬럼이 부채
            if cur_category == "금융자산" and safe(row, 0) == "대출":
                if v_amt_debt is not None:
                    try:
                        v_amt = -abs(float(v_amt_debt))
                    except (TypeError, ValueError):
                        pass

        try:
            amount = float(v_amt) if v_amt is not None else 0.0
        except (ValueError, TypeError):
            continue  # 금액이 숫자가 아니면 스킵

        if amount == 0:
            continue  # 0원 행은 계획 섹션 누출 또는 빈 행

        # ── 날짜 필드 ──
        start_date   = parse_date_str(safe(row, IDX_START))
        maturity_date = parse_date_str(safe(row, IDX_MATUR))
        payout_date  = parse_date_str(safe(row, IDX_PAYOUT))

        # ── 기타 필드 ──
        status       = str(safe(row, IDX_STATUS) or "").strip() or None
        product_name = str(safe(row, IDX_PROD) or "").strip() or None if IDX_PROD is not None else None
        memo         = str(safe(row, IDX_MEMO) or "").strip() or None if IDX_MEMO is not None else None

        # Layout B: status가 숫자(국민연금 금액)면 비우기
        if status and re.match(r"^\d+$", str(status)):
            status = None

        records.append({
            "category":          cur_category or "금융자산",
            "subcategory":       cur_subcategory or "보험·연금",
            "account_name":      str(v_name).strip(),
            "amount":            amount,
            "start_date":        start_date,
            "maturity_date":     maturity_date,
            "payout_start_date": payout_date,
            "status":            status,
            "product_name":      product_name,
            "memo":              memo,
            "owner":             infer_owner(str(v_name)),
        })

    return records


# ─────────────────────── DB 저장 ─────────────────────────────

def get_or_create_subcategory_id(cat_name: str, sub_name: str) -> int:
    cats = {c["name"]: c["id"] for c in repo.list_categories()}
    subs = repo.list_subcategories()

    # 카테고리 없으면 생성
    if cat_name not in cats:
        repo.add_category(cat_name, 99)
        cats = {c["name"]: c["id"] for c in repo.list_categories()}

    cat_id = cats[cat_name]

    # 서브카테고리 없으면 생성
    match = next((s for s in subs
                  if s["category_id"] == cat_id and s["name"] == sub_name), None)
    if match:
        return match["id"]

    repo.add_subcategory(cat_id, sub_name, 99)
    subs = repo.list_subcategories()
    return next(s["id"] for s in subs
                if s["category_id"] == cat_id and s["name"] == sub_name)


def get_or_create_owner_id(owner_name: str) -> int:
    owners = {o["name"]: o["id"] for o in repo.list_owners()}
    if owner_name not in owners:
        repo.add_owner(owner_name)
        owners = {o["name"]: o["id"] for o in repo.list_owners()}
    return owners[owner_name]


def find_or_create_account(record: dict) -> int:
    """
    (subcategory, account_name) 조합으로 기존 계좌를 찾고,
    없으면 새로 생성 후 id 반환.
    첫 발견 시 start_date, payout_start_date 등 메타정보도 채워준다.
    """
    sub_id   = get_or_create_subcategory_id(record["category"], record["subcategory"])
    owner_id = get_or_create_owner_id(record["owner"])

    existing_accounts = repo.list_accounts(active_only=False)
    match = next(
        (a for a in existing_accounts
         if a["subcategory_id"] == sub_id and a["name"] == record["account_name"]),
        None,
    )

    if match:
        # 메타 필드가 비어 있으면 보완
        needs_update = False
        update_fields = {
            "subcategory_id":    match["subcategory_id"],
            "owner_id":          match["owner_id"],
            "name":              match["name"],
            "product_name":      match["product_name"],
            "start_date":        match["start_date"],
            "maturity_date":     match["maturity_date"],
            "payout_start_date": match["payout_start_date"],
            "monthly_premium":   match["monthly_premium"],
            "expected_payout":   match["expected_payout"],
            "payout_type":       match["payout_type"],
            "status":            match["status"],
            "memo":              match["memo"],
            "is_active":         match["is_active"],
        }
        for field in ("start_date", "maturity_date", "payout_start_date"):
            if not update_fields[field] and record.get(field):
                update_fields[field] = record[field]
                needs_update = True
        if not update_fields["product_name"] and record.get("product_name"):
            update_fields["product_name"] = record["product_name"]
            needs_update = True
        if needs_update:
            repo.upsert_account(account_id=match["id"], **update_fields)
        return match["id"]

    # 신규 생성
    return repo.upsert_account(
        subcategory_id=sub_id,
        owner_id=owner_id,
        name=record["account_name"],
        product_name=record.get("product_name"),
        start_date=record.get("start_date"),
        maturity_date=record.get("maturity_date"),
        payout_start_date=record.get("payout_start_date"),
        monthly_premium=None,
        expected_payout=None,
        payout_type="monthly",
        status=record.get("status"),
        memo=record.get("memo"),
        is_active=1,
    )


def import_workbook(file_path_or_buffer) -> dict:
    """
    엑셀 파일 전체를 읽어서 금융현황_YYYYMM 시트를 모두 DB에 import.

    반환:
    {
      "sheets": [
        {
          "name": "금융현황_202601",
          "snapshot_date": "2026-01-01",
          "records": [...],   # 파싱된 원본
          "imported": int,    # 저장된 잔액 건수
          "skipped": int,     # 이미 존재해서 건너뜀
          "errors": [str]     # 오류 메시지
        }, ...
      ]
    }
    """
    wb = load_workbook(file_path_or_buffer, data_only=True)
    results = []

    target_sheets = sorted(
        [n for n in wb.sheetnames if re.match(r"금융현황_\d{6}$", n)],
        key=lambda n: n.split("_")[1],  # 연월 오름차순
    )

    for sheet_name in target_sheets:
        snap_date = sheet_name_to_date(sheet_name)
        ws = wb[sheet_name]

        sheet_result = {
            "name": sheet_name,
            "snapshot_date": snap_date,
            "records": [],
            "imported": 0,
            "skipped": 0,
            "errors": [],
        }

        if not snap_date:
            sheet_result["errors"].append("날짜 파싱 실패 - 스킵")
            results.append(sheet_result)
            continue

        try:
            records = parse_sheet(ws)
        except Exception as e:
            sheet_result["errors"].append(f"파싱 오류: {e}")
            results.append(sheet_result)
            continue

        sheet_result["records"] = records

        if not records:
            sheet_result["errors"].append("파싱된 항목 없음")
            results.append(sheet_result)
            continue

        # 스냅샷 생성 (이미 있으면 기존 것 사용)
        snap = repo.get_or_create_snapshot(snap_date)
        snap_id = snap["id"]

        # 이미 이 스냅샷에 잔액이 존재하면 → 기존 잔액 덮어쓸지 여부를 호출자가 결정
        # 여기서는 upsert(ON CONFLICT) 전략을 사용하므로 항상 최신값으로 갱신됨

        for rec in records:
            try:
                acc_id = find_or_create_account(rec)
                repo.save_balances(snap_id, [{"account_id": acc_id, "amount": rec["amount"]}])
                sheet_result["imported"] += 1
            except Exception as e:
                sheet_result["errors"].append(f"{rec.get('account_name', '?')}: {e}")

        results.append(sheet_result)

    return {"sheets": results}
